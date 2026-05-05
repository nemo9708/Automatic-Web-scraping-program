import os
import sys
import time
import smtplib
import base64
import requests
from io import BytesIO
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from datetime import datetime, timedelta
import builtins

# ==============================================================
# 📂 경로 동적 설정 (중요: 하드코딩 방지)
# ==============================================================
def get_base_path():
    """실행 파일(.exe) 또는 스크립트(.py)가 위치한 폴더 경로를 반환합니다."""
    if getattr(sys, 'frozen', False):
        # PyInstaller 등으로 빌드된 .exe 파일인 경우
        return os.path.dirname(sys.executable)
    # 일반 .py 스크립트 실행인 경우
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_path()

# GCP JSON 키 파일 경로 설정 (이제 내 컴퓨터 경로가 아닌 '현재 폴더'를 기준으로 찾음)
# 다른 사람에게 줄 때 이 JSON 파일을 프로그램과 같은 폴더에 두기만 하면 됩니다.
GCP_KEY_FILENAME = "qoo10-scraper-auto-22518a3e5c96.json"
GCP_KEY_PATH = os.path.join(BASE_DIR, GCP_KEY_FILENAME)

# ==============================================================
# 🕒 Timestamped Print
# ==============================================================
_original_print = builtins.print
def timestamped_print(*args, **kwargs):
    now = (datetime.utcnow() + timedelta(hours=9)).strftime("[%Y-%m-%d %H:%M:%S]")
    _original_print(now, *args, **kwargs)
builtins.print = timestamped_print

# ==============================================================
# 🔐 환경 변수 로드 (경로 문제 해결을 위해 환경 변수에서 경로를 읽도록 구성 가능)
# ==============================================================
QOO10_URL = os.getenv("QOO10_URL")
HIGHLIGHT_NAME1 = os.getenv("HIGHLIGHT_NAME1")
HIGHLIGHT_NAME2 = os.getenv("HIGHLIGHT_NAME2")
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_PASS = os.getenv("GMAIL_PASS")
SEND_TO = os.getenv("SEND_TO")

# 만약 스케줄러 온/오프 기능에서 경로 에러가 난다면, 
# 해당 기능을 수행하는 코드에서 반드시 위에서 정의한 'GCP_KEY_PATH'를 사용해야 합니다.

# ==============================================================
# 🖥 Chrome 설정
# ==============================================================
chrome_options = Options()
chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--window-size=1920,1080")
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
chrome_options.add_argument(f"--user-agent={user_agent}")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

# ==============================================================
# 📜 스크롤 다운 (중요: 이미지 Lazy Loading 해제)
# ==============================================================
def scroll_to_bottom(driver):
    print("[INFO] 페이지 스크롤 시작 (이미지를 깨우기 위해 천천히 스크롤)...")
    last_height = driver.execute_script("return document.body.scrollHeight")
    current_position = 0
    step = 800

    while current_position < last_height:
        current_position += step
        driver.execute_script(f"window.scrollTo(0, {current_position});")
        time.sleep(0.4)
        last_height = driver.execute_script("return document.body.scrollHeight")
        
    driver.execute_script("window.scrollTo(0, 0);")
    print("[INFO] 페이지 스크롤 완료")

# ==============================================================
# 🧩 메가와리 파서
# ==============================================================
def parse_megawari(driver):
    results = []
    items = driver.find_elements(By.CSS_SELECTOR, "ul.megasale_rank_list > li, ul.type_gallery > li, ul.list_item > li")
    print(f"[INFO] 감지된 아이템 수: {len(items)}")

    for item in items[:100]:
        try:
            rank = item.find_element(By.CSS_SELECTOR, ".rank_num").text.strip()
        except: rank = ""
        try:
            name = item.find_element(By.CSS_SELECTOR, ".title, .sbj").text.strip()
        except: name = ""
        try:
            total = item.find_element(By.CSS_SELECTOR, ".price, .prc strong").text.strip()
        except: total = ""
        try:
            img_el = item.find_element(By.CSS_SELECTOR, "img")
            img_url = (
                img_el.get_attribute("gd_src")
                or img_el.get_attribute("data-src")
                or img_el.get_attribute("data-original")
                or img_el.get_attribute("data-lazy")
                or img_el.get_attribute("src")
            )
        except: img_url = ""

        results.append([rank, name, total, img_url])
    return results

# ==============================================================
# 🚀 메인 로직 실행
# ==============================================================
try:
    print(f"[INFO] Qoo10 접속: {QOO10_URL}")
    if not QOO10_URL:
        raise ValueError("QOO10_URL 환경변수가 없습니다.")

    driver.get(QOO10_URL)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    scroll_to_bottom(driver)
    data = parse_megawari(driver)

    # 📘 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Qoo10 Ranking"
    ws.append(["순위", "상품명", "총판매액", "이미지"])

    for row in data:
        ws.append(row[:-1])

    # 🎨 하이라이트 적용
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_col=3):
        product_name = str(row[1].value) if row[1].value else ""
        should_highlight = False
        if HIGHLIGHT_NAME1 and HIGHLIGHT_NAME1 in product_name: should_highlight = True
        if HIGHLIGHT_NAME2 and HIGHLIGHT_NAME2 in product_name: should_highlight = True

        if should_highlight:
            for cell in row:
                cell.fill = yellow_fill
                cell.font = bold_font

    # 🖼 이미지 삽입 (브라우저 세션 연동)
    print("[INFO] 이미지 다운로드 시작...")
    session = requests.Session()
    session.headers.update({"User-Agent": user_agent, "Referer": driver.current_url})
    
    for cookie in driver.get_cookies():
        session.cookies.set(cookie['name'], cookie['value'])

    ws.column_dimensions['D'].width = 12

    for i, row in enumerate(data, start=2):
        img_url = row[3]
        if not img_url or img_url.startswith("data:image"): continue
        if img_url.startswith("//"): img_url = "https:" + img_url

        try:
            resp = session.get(img_url, timeout=10)
            if resp.status_code == 200:
                img_bytes = resp.content
                image = Image.open(BytesIO(img_bytes))
                image = image.convert("RGB")
                image.thumbnail((80, 80))
                bio = BytesIO()
                image.save(bio, format="PNG")
                bio.seek(0)
                img = XLImage(bio)
                ws.row_dimensions[i].height = 65 
                ws.add_image(img, f"D{i}")
                time.sleep(0.1)
        except Exception as e:
            print(f"[ERROR] 이미지 삽입 실패 (순위 {row[rank]}): {e}")

    # 💾 저장 및 메일 전송
    # 저장 경로 역시 실행 위치(BASE_DIR)를 기준으로 설정하면 더 안전합니다.
    file_path = os.path.join(BASE_DIR, "Qoo10_Rank.xlsx")
    wb.save(file_path)
    print(f"[INFO] 엑셀 저장 완료: {file_path}")

    if GMAIL_USER and GMAIL_PASS and SEND_TO:
        msg = MIMEMultipart()
        msg["From"] = GMAIL_USER
        msg["To"] = SEND_TO
        today = (datetime.utcnow() + timedelta(hours=9)).strftime("%Y-%m-%d")
        msg["Subject"] = f"Qoo10 랭킹 자동 보고서 {today}"
        
        body = MIMEText(f"자동 생성된 Qoo10 보고서입니다.\n일자: {today}\nURL: {QOO10_URL}", "plain")
        msg.attach(body)

        with open(file_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename=Qoo10_Rank.xlsx")
        msg.attach(part)

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(GMAIL_USER, GMAIL_PASS)
            server.send_message(msg)
        print("[INFO] 메일 전송 완료")
    
except Exception as e:
    print(f"[ERROR] 프로그램 실행 중 오류: {e}")

finally:
    if 'driver' in locals():
        driver.quit()
        print("[INFO] 브라우저 종료")